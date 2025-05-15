from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import datetime

def create_excel_file(filename, sheet_name, headers, data):
    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Add headers
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    # Add data
    for row, record in enumerate(data, start=2):
        for col, value in enumerate(record, start=1):
            ws.cell(row=row, column=col, value=value)

    # Adjust column widths
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 15

    # Save the workbook
    wb.save(filename)

    print(f"Excel file '{filename}' has been created successfully.")
